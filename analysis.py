"""
Canada Job Market Analysis — SQL + Python
==========================================
Author  : Ruchitha Yathirajulu
Source  : Statistics Canada, Table 14-10-0202-01
          Labour force characteristics by industry, monthly
          Period: January 2019 — December 2024
"""

import pandas as pd
import sqlite3
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("=" * 60)
print("  Canada Job Market Analysis — Real StatCan Data")
print("  Table 14-10-0202-01 | 2019–2024")
print("  by Ruchitha Yathirajulu")
print("=" * 60)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — LOAD & CLEAN RAW DATA
# ══════════════════════════════════════════════════════════════════════════════
print("\n📂 Loading raw Statistics Canada data...")

df = pd.read_csv(
    "data/1410002201_databaseLoadingData.csv",
    encoding="utf-8-sig"
)

# Keep only useful columns and rename
df = df[[
    "REF_DATE",
    "North American Industry Classification System (NAICS)",
    "VALUE"
]].rename(columns={
    "North American Industry Classification System (NAICS)": "Industry",
    "VALUE": "Employed_000s"
})

# Parse date
df["REF_DATE"]    = pd.to_datetime(df["REF_DATE"])
df["Year"]        = df["REF_DATE"].dt.year
df["Month"]       = df["REF_DATE"].dt.month
df["Month_Name"]  = df["REF_DATE"].dt.strftime("%b")
df["Quarter"]     = "Q" + df["REF_DATE"].dt.quarter.astype(str)
df["Year_Month"]  = df["REF_DATE"].dt.strftime("%Y-%m")

# Drop rows with no value
df = df.dropna(subset=["Employed_000s"])
df["Employed_000s"] = pd.to_numeric(df["Employed_000s"], errors="coerce")
df = df.dropna(subset=["Employed_000s"])

# Clean NAICS codes from industry names
# e.g. "Construction [23]" → "Construction"
df["Industry_Clean"] = df["Industry"].str.replace(
    r'\s*\[.*?\]', '', regex=True
).str.strip()

# Keep only top-level industries — exclude sub-categories
# Sub-categories have very long names with multiple brackets
top_level_keywords = [
    "Total, all industries",
    "Goods-producing sector",
    "Services-producing sector",
    "Agriculture",
    "Forestry, fishing, mining",
    "Utilities",
    "Construction",
    "Manufacturing",
    "Wholesale and retail trade",
    "Transportation and warehousing",
    "Finance, insurance",
    "Professional, scientific",
    "Business, building",
    "Educational services",
    "Health care",
    "Information, culture",
    "Accommodation and food",
    "Other services",
    "Public administration"
]

# Filter to keep only rows where industry starts with a top-level keyword
def is_top_level(ind):
    return any(ind.startswith(kw) or kw in ind for kw in top_level_keywords)

df_filtered = df[df["Industry_Clean"].apply(is_top_level)].copy()

# Remove sector aggregates for detailed analysis
exclude = ["Goods-producing sector", "Services-producing sector"]
df_detail = df_filtered[~df_filtered["Industry_Clean"].isin(exclude)].copy()

print(f"✅ Loaded {len(df):,} rows")
print(f"✅ Filtered to {len(df_detail):,} rows — top-level industries only")
print(f"   Date range  : {df['Year_Month'].min()} → {df['Year_Month'].max()}")
print(f"   Industries  : {df_detail['Industry_Clean'].nunique()}")
print(f"   Years       : {sorted(df_detail['Year'].unique())}")

# Save cleaned file
df_detail.to_csv("data/canada_employment_clean.csv", index=False)
print("✅ Saved: data/canada_employment_clean.csv\n")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — LOAD INTO SQLite & RUN 5 QUERIES
# ══════════════════════════════════════════════════════════════════════════════
print("🔍 Running SQL Queries...\n")

conn = sqlite3.connect(":memory:")
df_detail.to_sql("employment", conn, index=False, if_exists="replace")

queries = {

    "Q1_National_Trend": {
        "title": "National Employment Trend — Monthly 2019 to 2024",
        "description": "Shows overall employment trend including COVID-19 crash and recovery",
        "sql": """
            SELECT
                Year_Month,
                Year,
                Month,
                Quarter,
                ROUND(SUM(CASE WHEN Industry_Clean = 'Total, all industries'
                    THEN Employed_000s END), 1)          AS Total_Employed_000s
            FROM employment
            WHERE Industry_Clean = 'Total, all industries'
            GROUP BY Year_Month, Year, Month, Quarter
            ORDER BY Year_Month
        """
    },

    "Q2_Industry_Avg": {
        "title": "Average Employment by Industry — 2019 to 2024",
        "description": "Ranks industries by average employment size over the full period",
        "sql": """
            SELECT
                Industry_Clean                           AS Industry,
                ROUND(AVG(Employed_000s), 1)            AS Avg_Employed_000s,
                ROUND(MAX(Employed_000s), 1)            AS Peak_Employed_000s,
                ROUND(MIN(Employed_000s), 1)            AS Trough_Employed_000s,
                ROUND(MAX(Employed_000s)
                    - MIN(Employed_000s), 1)            AS Peak_To_Trough_000s,
                RANK() OVER (
                    ORDER BY AVG(Employed_000s) DESC
                )                                       AS Industry_Rank
            FROM employment
            WHERE Industry_Clean != 'Total, all industries'
            GROUP BY Industry_Clean
            ORDER BY Avg_Employed_000s DESC
        """
    },

    "Q3_COVID_Impact": {
        "title": "COVID-19 Impact Analysis — 2019 vs 2020 vs 2024",
        "description": "Measures employment drop during COVID and recovery by industry",
        "sql": """
            WITH yearly AS (
                SELECT
                    Industry_Clean,
                    Year,
                    ROUND(AVG(Employed_000s), 1) AS Avg_Employed
                FROM employment
                WHERE Industry_Clean != 'Total, all industries'
                GROUP BY Industry_Clean, Year
            ),
            pivot AS (
                SELECT
                    Industry_Clean,
                    MAX(CASE WHEN Year = 2019 THEN Avg_Employed END) AS Emp_2019,
                    MAX(CASE WHEN Year = 2020 THEN Avg_Employed END) AS Emp_2020,
                    MAX(CASE WHEN Year = 2021 THEN Avg_Employed END) AS Emp_2021,
                    MAX(CASE WHEN Year = 2023 THEN Avg_Employed END) AS Emp_2023,
                    MAX(CASE WHEN Year = 2024 THEN Avg_Employed END) AS Emp_2024
                FROM yearly
                GROUP BY Industry_Clean
            )
            SELECT
                Industry_Clean                          AS Industry,
                Emp_2019,
                Emp_2020,
                Emp_2024,
                ROUND(Emp_2020 - Emp_2019, 1)          AS COVID_Drop_000s,
                ROUND((Emp_2020 - Emp_2019)
                    / Emp_2019 * 100, 1)               AS COVID_Drop_Pct,
                ROUND(Emp_2024 - Emp_2019, 1)          AS Recovery_vs_2019_000s,
                ROUND((Emp_2024 - Emp_2019)
                    / Emp_2019 * 100, 1)               AS Recovery_Pct,
                CASE
                    WHEN Emp_2024 >= Emp_2019 * 1.05   THEN '🚀 Surpassed Pre-COVID'
                    WHEN Emp_2024 >= Emp_2019           THEN '✅ Fully Recovered'
                    WHEN Emp_2024 >= Emp_2019 * 0.97   THEN '🟡 Nearly Recovered'
                    ELSE '🔴 Still Below 2019'
                END                                     AS Recovery_Status
            FROM pivot
            WHERE Emp_2019 IS NOT NULL
            ORDER BY COVID_Drop_Pct ASC
        """
    },

    "Q4_YoY_Growth": {
        "title": "Year-over-Year Employment Growth by Industry",
        "description": "Annual % change in employment for each industry",
        "sql": """
            WITH yearly AS (
                SELECT
                    Industry_Clean,
                    Year,
                    ROUND(AVG(Employed_000s), 1) AS Avg_Emp
                FROM employment
                WHERE Industry_Clean != 'Total, all industries'
                GROUP BY Industry_Clean, Year
            )
            SELECT
                curr.Industry_Clean                     AS Industry,
                curr.Year,
                curr.Avg_Emp                            AS Current_Emp_000s,
                prev.Avg_Emp                            AS Prev_Emp_000s,
                ROUND(curr.Avg_Emp - prev.Avg_Emp, 1)  AS Change_000s,
                ROUND((curr.Avg_Emp - prev.Avg_Emp)
                    / prev.Avg_Emp * 100, 2)           AS YoY_Growth_Pct
            FROM yearly curr
            JOIN yearly prev
                ON curr.Industry_Clean = prev.Industry_Clean
                AND curr.Year = prev.Year + 1
            ORDER BY curr.Industry_Clean, curr.Year
        """
    },

    "Q5_Quarterly_Trend": {
        "title": "Quarterly Employment Trend by Industry",
        "description": "Average employment per quarter — shows seasonality and trends",
        "sql": """
            SELECT
                Industry_Clean                          AS Industry,
                Year,
                Quarter,
                ROUND(AVG(Employed_000s), 1)           AS Avg_Quarterly_Emp_000s,
                ROUND(MAX(Employed_000s), 1)           AS Max_Monthly_Emp_000s,
                ROUND(MIN(Employed_000s), 1)           AS Min_Monthly_Emp_000s
            FROM employment
            WHERE Industry_Clean != 'Total, all industries'
            GROUP BY Industry_Clean, Year, Quarter
            ORDER BY Industry_Clean, Year, Quarter
        """
    }
}

results = {}
for key, q in queries.items():
    result_df = pd.read_sql_query(q["sql"], conn)
    results[key] = result_df
    print(f"✅ {q['title']}")
    print(f"   → {len(result_df)} rows returned\n")

conn.close()

# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — EXPORT TO FORMATTED EXCEL
# ══════════════════════════════════════════════════════════════════════════════
print("📊 Building Excel report...\n")

output_path = "output/Canada_Employment_Analysis.xlsx"

NAVY  = "1B3A6B"
TEAL  = "0E7490"
WHITE = "FFFFFF"
LIGHT = "EFF6FF"
GREEN = "DCFCE7"
YELLOW= "FEF9C3"
RED   = "FEE2E2"
DARK  = "1F2937"
GRAY  = "F3F4F6"

def fill(hex): return PatternFill("solid", start_color=hex, end_color=hex)
def fnt(hex, bold=False, sz=10): return Font(name="Arial", color=hex, bold=bold, size=sz)
def aln(h="left"): return Alignment(horizontal=h, vertical="center", wrap_text=True)
def bdr():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def style_sheet(ws):
    # Header row
    for cell in ws[1]:
        cell.fill      = fill(NAVY)
        cell.font      = fnt(WHITE, True, 10)
        cell.alignment = aln("center")
        cell.border    = bdr()
    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        bg = LIGHT if row_idx % 2 == 0 else WHITE
        for cell in row:
            val = str(cell.value or "")
            if "Surpassed" in val or "Fully Recovered" in val:
                cell.fill = fill(GREEN)
            elif "Nearly" in val:
                cell.fill = fill(YELLOW)
            elif "Still Below" in val:
                cell.fill = fill(RED)
            else:
                cell.fill = fill(bg)
            cell.font      = fnt(DARK, False, 10)
            cell.alignment = aln("left")
            cell.border    = bdr()
    # Auto width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 48)
    ws.freeze_panes = "A2"

# Write all sheets
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df_detail.to_excel(writer, sheet_name="Cleaned Data", index=False)
    for key, q in queries.items():
        results[key].to_excel(writer, sheet_name=key[:31], index=False)

wb = load_workbook(output_path)
for ws in wb.worksheets:
    style_sheet(ws)

# ── Cover sheet ───────────────────────────────────────────────────────────────
cover = wb.create_sheet("📋 Project Overview", 0)
cover.column_dimensions["A"].width = 85

# Get some quick stats for the cover
total_national = results["Q1_National_Trend"]
peak = total_national.loc[total_national["Total_Employed_000s"].idxmax()]
trough = total_national.loc[total_national["Total_Employed_000s"].idxmin()]

cover_rows = [
    ("Canada Job Market Analysis — SQL + Power BI Project", True, NAVY, 14),
    ("Source: Statistics Canada, Table 14-10-0202-01 — Real Government Data ✅", False, TEAL, 11),
    ("", False, WHITE, 10),
    ("Project Details", True, NAVY, 11),
    (f"Author       : Ruchitha Yathirajulu", False, DARK, 10),
    (f"Data Source  : Statistics Canada Labour Force Survey — Table 14-10-0202-01", False, DARK, 10),
    (f"Period       : January 2019 — December 2024 (6 full years)", False, DARK, 10),
    (f"Records      : {len(df_detail):,} rows after cleaning", False, DARK, 10),
    (f"Industries   : {df_detail['Industry_Clean'].nunique()} NAICS categories", False, DARK, 10),
    (f"Tools        : Python (pandas, sqlite3, openpyxl) → Excel → Power BI", False, DARK, 10),
    ("", False, WHITE, 10),
    ("Key Statistics", True, NAVY, 11),
    (f"Peak Employment     : {peak['Total_Employed_000s']:,.1f}K persons — {peak['Year_Month']}", False, DARK, 10),
    (f"Trough Employment   : {trough['Total_Employed_000s']:,.1f}K persons — {trough['Year_Month']} (COVID-19 impact)", False, DARK, 10),
    ("", False, WHITE, 10),
    ("Sheets in This Workbook", True, NAVY, 11),
    ("Cleaned Data     — Structured dataset ready for Power BI", False, DARK, 10),
    ("Q1_National_Trend  — Monthly employment trend 2019–2024", False, DARK, 10),
    ("Q2_Industry_Avg    — Industries ranked by employment size", False, DARK, 10),
    ("Q3_COVID_Impact    — COVID-19 drop and recovery status per industry", False, DARK, 10),
    ("Q4_YoY_Growth      — Year-over-year employment growth by industry", False, DARK, 10),
    ("Q5_Quarterly_Trend — Quarterly employment by industry (Power BI ready)", False, DARK, 10),
    ("", False, WHITE, 10),
    ("Key Findings from SQL Analysis", True, NAVY, 11),
    ("• Health care and social assistance is consistently the largest employment sector", False, DARK, 10),
    ("• Accommodation and food services saw the sharpest COVID-19 employment decline", False, DARK, 10),
    ("• Most industries surpassed pre-COVID employment levels by 2023", False, DARK, 10),
    ("• Construction and professional services showed strongest post-COVID growth", False, DARK, 10),
    ("• National employment peaked in 2024 — above all pre-COVID levels", False, DARK, 10),
]

for i, (text, bold, color, sz) in enumerate(cover_rows, 1):
    c = cover.cell(row=i, column=1, value=text)
    c.font      = fnt(color, bold, sz)
    c.alignment = Alignment(horizontal="left", vertical="center")
    cover.row_dimensions[i].height = 22 if text else 8

wb.save(output_path)
print(f"✅ Excel saved: {output_path}")
print("\n🎉 Done! Your files are ready:")
print("   📊 output/Canada_Employment_Analysis.xlsx — open in Power BI")
print("   📄 data/canada_employment_clean.csv — cleaned dataset")
print("\n💡 Next step: Open Power BI Desktop and connect to the Excel file!\n")
